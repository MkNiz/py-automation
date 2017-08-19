import openpyxl

book = openpyxl.Workbook()
sheet = book.get_active_sheet()

sheet['A1'] = "Numbers that are digestible by human stomachs"
# Set column width
sheet.column_dimensions['A'].width = 8
# Populate column's cells
for i in range(2, 101):
    n = i - 1
    sheet['A' + str(i)] = int(n * (n / 2))
    
# Set row height for every other row with data
for i in range(2, 101, 2):
    sheet.row_dimensions[i].height = 18

# Merge cells to expand header column
sheet.merge_cells('A1:E1')
# Demonstrating unmerge before merging back
sheet.unmerge_cells('A1:E1')
sheet.merge_cells('A1:E1')

# Freeze the header
sheet.freeze_panes = 'A2'

book.save("edible_numbers.xlsx")
print("Saved spreadsheet to edible_numbers.xlsx")